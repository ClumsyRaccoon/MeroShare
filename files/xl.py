import os
import pandas as pd
from openpyxl import load_workbook
import datetime

try:
    from meroshare import MeroShare
except:
    from files.meroshare import MeroShare


def get_login_info(details, client_type):
    return {
        "name": details[1],
        "username": details[4].replace(" ", ""),
        "password": details[6],
        "dpid": int(details[5]) - 13000000,
        "client_id": client_type + str(details[0]),
        "crn": details[7] if len(details) > 7 else None,
        "pin": details[8] if len(details) > 8 else None,
        "bank": details[9] if len(details) > 9 else None,
    }


def apply_ipo(sheet, full_list, client_type, Scrip, qty):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO" or str(details[3]).upper() == "NO":
            continue

        login_info = get_login_info(details, client_type)
        ms = MeroShare(**login_info)
        if ms.login():
            ms.apply(Scrip, qty)
        data = [
            ms.client_id,
            details[1],
            str(details[5]) + details[4].replace(" ", ""),
            Scrip,
            ms.status,
        ]
        full_list.loc[len(full_list)] = data
    return full_list


def check_account_status(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO":
            continue

        login_info = get_login_info(details, client_type)
        ms = MeroShare(**login_info)
        ms.login()
        data = [
            ms.client_id,
            details[1],
            str(details[5]) + details[4].replace(" ", ""),
            ms.status,
        ]
        full_list.loc[len(full_list)] = data
    return full_list


def get_applicable_issues(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO" or str(details[3]).upper() == "NO":
            continue

        login_info = get_login_info(details, client_type)
        ms = MeroShare(**login_info)
        if ms.login():
            try:
                for item in ms.get_applicable_issues():
                    data = [
                        ms.client_id,
                        details[1],
                        str(details[5]) + details[4].replace(" ", ""),
                        item["scrip"],
                        item["shareGroupName"],
                        item["shareTypeName"],
                        item.get("reservationTypeName", "NA"),
                    ]
                    full_list.loc[len(full_list)] = data
            except:
                full_list.loc[len(full_list)] = [
                    ms.client_id,
                    details[1],
                    str(details[5]) + details[4].replace(" ", ""),
                    ms.status,
                    "NA",
                    "NA",
                    "NA",
                ]
        else:
            full_list.loc[len(full_list)] = [
                ms.client_id,
                details[1],
                str(details[5]) + details[4].replace(" ", ""),
                ms.status,
                "NA",
                "NA",
                "NA",
            ]
    return full_list


def check_ipo_status(sheet, full_list, client_type, Scrip):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO" or str(details[3]).upper() == "NO":
            continue

        login_info = get_login_info(details, client_type)
        ms = MeroShare(**login_info)
        ms.login()
        ms.get_application_status(Scrip)
        data = [
            ms.client_id,
            details[1],
            str(details[5]) + details[4].replace(" ", ""),
            Scrip,
            ms.status,
        ]
        full_list.loc[len(full_list)] = data
    return full_list


def list_shares(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO":
            continue

        login_info = get_login_info(details, client_type)
        ms = MeroShare(**login_info)
        if ms.login():
            try:
                full_list = (
                    ms.get_share_list()
                    if full_list.empty
                    else pd.concat([full_list, ms.get_share_list()])
                )
            except:
                full_list.loc[len(full_list)] = [
                    client_type + str(details[0]),
                    details[1],
                    str(details[5]) + details[4].replace(" ", ""),
                    ms.status,
                    0,
                    0,
                ]
        else:
            full_list.loc[len(full_list)] = [
                client_type + str(details[0]),
                details[1],
                str(details[5]) + details[4].replace(" ", ""),
                ms.status,
                0,
                0,
            ]
    return full_list


def main():
    book = load_workbook(filename="MeroShare Login Details.xlsx", data_only=True)
    sheet = book["List"]

    while True:
        print("Please select an option: \n")
        print("1. Check Account Status")
        print("2. List My Shares")
        print("3. Get Applicable Issues")
        print("4. Apply IPO")
        print("5. Check IPO Application Status")
        print("0. EXIT \n")

        choice = input("Enter your choice: ")

        if choice == "0":
            break

        elif choice == "1":
            df = pd.DataFrame(columns=["Client ID", "Name", "Demat", "Status"])
            df = check_account_status(sheet, df, "")
            df.to_excel("MeroShare Account Status.xlsx", index=False)

        elif choice == "2":
            df = pd.DataFrame(
                columns=[
                    "Client ID",
                    "Name",
                    "DMAT No",
                    "Script",
                    "Current Balance",
                    "Free Balance",
                ]
            )
            df = list_shares(sheet, df, "")
            filename = f'MeroShare - Share List - {datetime.datetime.now().strftime("%d-%b-%Y")}.xlsx'
            df.to_excel(filename, index=False)

        elif choice == "3":
            df = pd.DataFrame(
                columns=[
                    "Client ID",
                    "Name",
                    "Demat",
                    "Script",
                    "Share Group",
                    "Type",
                    "Reservation Type",
                ]
            )
            df = get_applicable_issues(sheet, df, "")
            df.to_excel("Applicable Issue List.xlsx", index=False)

        elif choice == "4":
            scrip = input("Script Code to Apply For: ")
            qty = input("No. of Kitta to Apply: ")
            df = pd.DataFrame(
                columns=["Client ID", "Name", "Demat", "Script", "Application"]
            )
            df = apply_ipo(sheet, df, "", scrip, qty)
            df.to_excel(f"IPO Applied for {scrip}.xlsx", index=False)

        elif choice == "5":
            scrip = input("Script Code to Check: ")
            df = pd.DataFrame(columns=["Client ID", "Name", "Demat", "Scrip", "Status"])
            df = check_ipo_status(sheet, df, "", scrip)
            df.to_excel(f"Application Status for {scrip}.xlsx", index=False)

        else:
            print("Invalid choice!")
            continue

        input("Press Enter to Continue....")

        os.system("cls")


if __name__ == "__main__":
    main()
