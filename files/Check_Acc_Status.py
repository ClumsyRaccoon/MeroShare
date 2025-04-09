import pandas as pd
from openpyxl import load_workbook

try:
    from meroshare import MeroShare
except:
    from files.meroshare import MeroShare


def application_list(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):

        if str(details[2]).upper() == "NO":
            continue

        login_info = {
            "name": details[1],
            "username": details[4].replace(" ", ""),
            "password": details[6],
            "dpid": int(details[5]) - 13000000,
            "client_id": client_type + str(details[0]),
            "crn": details[7],
            "pin": details[8],
        }

        ms = MeroShare(**login_info)
        login = ms.login()

        data = (
            [ms.client_id]
            + [details[1]]
            + [str(details[5]) + str(details[4].replace(" ", ""))]
            + [ms.status]
        )
        full_list.loc[len(full_list)] = data
        print("\n")
    return full_list


def read_excel():
    full_list = pd.DataFrame(columns=["Client ID", "Name", "Demat", "Status"])

    book = load_workbook(filename="MeroShare Login Details.xlsx", data_only=True)

    full_list = application_list(book["List"], full_list, "")

    print(full_list)
    full_list.to_excel(f"MeroShare Account Status.xlsx", index=False)


def start():
    read_excel()
    input("Press Enter to Continue....")


if __name__ == "__main__":
    start()
