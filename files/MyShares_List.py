import pandas as pd
from openpyxl import load_workbook
import datetime

try:
    from meroshare import MeroShare
except:
    from files.meroshare import MeroShare


def check(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(details[2]).upper() == "NO":
            continue

        login_info = {
            "name": details[1],
            "username": details[4].replace(" ", ""),
            "password": details[6],
            "dpid": int(details[5]) - 13000000,
            "client_id": client_type + str(details[0]),
        }

        ms = MeroShare(**login_info)
        login = ms.login()

        if login:
            try:
                full_list = (
                    ms.get_share_list()
                    if full_list.empty
                    else pd.concat([full_list, ms.get_share_list()])
                )
            except:
                data = (
                    [client_type + str(details[0])]
                    + [details[1]]
                    + [str(details[5]) + str(details[4].replace(" ", ""))]
                    + [ms.status]
                    + [0, 0]
                )
                full_list.loc[len(full_list)] = data
        else:
            data = (
                [client_type + str(details[0])]
                + [details[1]]
                + [str(details[5]) + str(details[4].replace(" ", ""))]
                + [ms.status]
                + [0, 0]
            )
            full_list.loc[len(full_list)] = data

    print("\n")

    return full_list


def check_share():
    full_list = pd.DataFrame(
        columns=[
            "Client ID",
            "Name",
            "DMAT No",
            "Script",
            "Current Balance",
            "Free Balance",
        ]
    )

    book = load_workbook(filename="MeroShare Login Details.xlsx", data_only=True)

    full_list = check(book["List"], full_list, "")

    print(full_list)
    full_list.to_excel(
        f'MeroShare - Share List - {datetime.datetime.now().strftime("%d-%b-%Y")}.xlsx',
        index=False,
    )

    input("Press Enter to Continue....")


if __name__ == "__main__":
    check_share()
