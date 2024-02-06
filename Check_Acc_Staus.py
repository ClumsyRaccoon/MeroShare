import logging
from tenacity import retry, stop_after_attempt, wait_fixed
import requests
import json
import os
import pandas as pd
from openpyxl import load_workbook

USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.105 Safari/537.36"
logging.basicConfig(format="%(asctime)s %(message)s", level=logging.INFO)


class MeroShare:
    def __init__(
            self,
            name: str = None,
            dpid: int = None,
            username: int = None,
            password: str = None,
            client_id: str = None,
            crn: str = None,
            pin: int = None
    ):

        self.__name = name
        self.__dpid = str(dpid)
        self.__username = str(username)
        self.__password = password
        self.__session = requests.Session()
        self.__auth_token = None
        self.status = None
        self.__capital_id = self.get_capital_id()
        self.__dmat = "130" + self.__dpid + self.__username
        self.client_id = client_id
        self.__crn = crn
        self.__pin = pin
        self.__applicable_issues = None
        self.__account = None


    def get_capital_id(self):
        if os.path.exists('capitals.json'):
            try:
                return \
                    [response['id'] for response in json.load(open('capitals.json')) if
                     response['code'] == self.__dpid][0]
            except Exception:
                print('Could not find capital id cache \n Updating cache and Retrying...')
        with self.__session as sess:
            headers = {
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "en-US,en;q=0.9",
                "Authorization": "null",
                "Connection": "keep-alive",
                "Origin": "https://meroshare.cdsc.com.np",
                "Referer": "https://meroshare.cdsc.com.np/",
                "Sec-Fetch-Dest": "empty",
                "Sec-Fetch-Mode": "cors",
                "Sec-Fetch-Site": "same-site",
                "Sec-GPC": "1",
                "User-Agent": USER_AGENT,
            }
            sess.headers.update(headers)

            try:
                cap_list = sess.get("https://webbackend.cdsc.com.np/api/meroShare/capital/").json()
                with open("capitals.json", "w") as cap_file:
                    json.dump(cap_list, cap_file)

                return [response['id'] for response in cap_list if response['code'] == self.__dpid][0]
            except Exception as error:
                self.status = 'Error finding capital id'
                logging.warning(f'{self.status} for Acc: {self.__name}')
                logging.info(error)
                return None

    @retry(stop=stop_after_attempt(3), wait=wait_fixed(3), reraise=True)
    def login(self) -> bool:
        assert (
                self.__username and self.__password and self.__dpid
        ), "Username, password and DPID required!"

        if not self.__capital_id:
            logging.warning("Capital ID Problem")
            return False

        try:
            with self.__session as sess:
                data = json.dumps(
                    {
                        "clientId": self.__capital_id,
                        "username": self.__username,
                        "password": self.__password,
                    }
                )

                headers = {
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Language": "en-US,en;q=0.9",
                    "Authorization": "null",
                    "Connection": "keep-alive",
                    "Content-Type": "application/json",
                    "Origin": "https://meroshare.cdsc.com.np",
                    "Referer": "https://meroshare.cdsc.com.np/",
                    "Sec-Fetch-Dest": "empty",
                    "Sec-Fetch-Mode": "cors",
                    "Sec-Fetch-Site": "same-site",
                    "Sec-GPC": "1",
                    "User-Agent": USER_AGENT,
                }
                sess.headers.update(headers)
                try:
                    sess.options("https://webbackend.cdsc.com.np/api/meroShare/auth/")
                    login_req = sess.post(
                        "https://webbackend.cdsc.com.np/api/meroShare/auth/", data=data)
                    self.status = login_req.json()['message']
                    if login_req.status_code == 200:
                        self.__auth_token = login_req.headers.get("Authorization")
                        
                        logging.info(f"{self.status}  Account: {self.__name}!")
                    else:
                        self.status = f'Login Failed! {login_req.status_code}'
                        logging.warning(f"{self.status} Account: {self.__name}")
                        print(f'Login Failed {self.__name}')
                except:
                    print('Login Connection Refused')
                    self.status = 'Login Connection Refused'

        except Exception as error:
            print(error)
            logging.info(error)
            logging.error(
                f"Login request failed! Retrying ({self.login.retry.statistics.get('attempt_number')})!"
            )
            self.status = 'Login Failed'
            return False

        return True


def application_list(sheet, full_list, client_type):
    for details in sheet.iter_rows(min_row=2,
                                   min_col=1,
                                   values_only=True):
        if str(details[2]).upper()=='NO' and str(details[3]).upper()=='NO':
            continue

        login_info = {"name": details[1],
                      "username": details[4].replace(" ", ""),
                      "password": details[6],
                      "dpid": int(details[5]) - 13000000,
                      "client_id": client_type + str(details[0]),
                      "crn": details[7],
                      "pin": details[8]}

        ms = MeroShare(**login_info)
        ms.login()

        data = [ms.client_id] + [details[1]] + [str(details[5]) + str(details[4].replace(" ",""))] + [ms.status]
        full_list.loc[len(full_list)] = data

        print('\n')

    return (full_list)


def  read_excel():
    full_list = pd.DataFrame(columns=['Client ID', 'Name', 'Demat', 'Status'])

    book = load_workbook(filename='MeroShare Login Details.xlsx', data_only=True)

    full_list = application_list(book['PMS List'], full_list, 'PMS ')
    full_list = application_list(book['Investment'], full_list, 'Inv ')
    full_list = application_list(book['Others'], full_list, 'Others ')

    print(full_list)
    full_list.to_excel(f'MeroShare Account Status.xlsx', index=False)


def start():
    read_excel()
    input('Press Enter to Continue....')


if __name__ == '__main__':
    start()
